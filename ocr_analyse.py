import cv2
import pytesseract
import re
import os
import pandas as pd
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

NAME_DIR="pliki"
BASE_DIR=os.path.dirname(os.path.abspath(__file__))
FILES_PATH=os.path.join(BASE_DIR,NAME_DIR)
CONFIG = r'--psm 11 --oem 3 -c tessedit_char_whitelist=0123456789,. -c classify_bln_numeric_mode=1'
sys.stdout.reconfigure(encoding="utf-8")
pytesseract.pytesseract.tesseract_cmd=r"C:\Program Files\Tesseract-OCR\tesseract.exe"
list_of_optical_measurements=[]


def preprocess_image(path):
    img = cv2.imread(path)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray = cv2.resize(gray, None, fx=2, fy=2,interpolation=cv2.INTER_CUBIC)
    return gray


def extract_numbers(img):
    text=pytesseract.image_to_string(img,config=CONFIG)
    text=text.replace(",", ".")
    raw_numbers=re.findall(r"\d+\.\d{3}",text)
    return raw_numbers


files=sorted(
    [f for f in os.listdir(FILES_PATH) if f.lower().endswith(".bmp")],
    key=lambda x:int(os.path.splitext(x)[0])
)



for file in files:
    path=os.path.join(FILES_PATH,file)
    img=preprocess_image(path)
    numbers=extract_numbers(img)
    if numbers:
        list_of_optical_measurements.append(numbers[:10])

print(list_of_optical_measurements)
if not list_of_optical_measurements:
    print("OCR nie znalazł danych")
    exit()


data=pd.DataFrame(list_of_optical_measurements)
data.index=range(1,len(data)+1)
data.columns=range(1,data.shape[1]+1)
output_path=os.path.join(BASE_DIR,"results.xlsx")
data.to_excel(output_path)


wb=load_workbook(output_path)
ws=wb.active

max_value=float(pd.to_numeric(data.stack(),errors="coerce").max())
min_value=float(pd.to_numeric(data.stack(),errors="coerce").min())


red_fill=PatternFill(start_color="FF0000",end_color="FF0000",fill_type="solid")
yellow_fill=PatternFill(start_color="FFFF00",end_color="FFFF00",fill_type="solid")


thin=Side(style="thin")
border=Border(left=thin,right=thin,top=thin,bottom=thin)

for row in ws.iter_rows(min_row=2,min_col=2,max_col=data.shape[1]+1):
    for cell in row:
        cell.border=border
        try:
            value=float(cell.value)
            if value==max_value:
                cell.fill=red_fill
            if value==min_value:
                cell.fill=yellow_fill
        except:
            pass

wb.save(output_path)

result = max_value / min_value
if result<2:
    print("TEST RESULT: Positive")
else:
    print("TEST RESULT: Negative")
print("Excel file was saved:",output_path)